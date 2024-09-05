import pandas as pd
from typing import List, Dict
from Diet_class import Ingredient, Menu, Meal, Diet, NutrientConstraints
from openpyxl import load_workbook

def load_and_process_data(diet_db_path: str, menu_db_path: str, ingre_db_path: str) -> Diet:
    diet_df = pd.read_excel(diet_db_path, sheet_name='sample')
    menu_ingre_df = pd.read_excel(menu_db_path, sheet_name='ingredient')
    menu_nutri_df = pd.read_excel(menu_db_path, sheet_name='nutrient')
    menu_cat_df = pd.read_excel(menu_db_path, sheet_name='category')
    ingre_price_df = pd.read_excel(ingre_db_path)

    menu_categories = dict(zip(menu_cat_df['Menu'], menu_cat_df['Category']))

    # ingredient_dict 생성
    ingredient_dict = {}
    for _, row in menu_ingre_df.iterrows():
        menu_name = row['Menu']
        ingredient_name = row['Ingredient']
        amount = row['Amount_g']
        
        price_per_100g = ingre_price_df[ingre_price_df['Ingredient'] == ingredient_name]['Price']
        price = (price_per_100g.values[0] / 100) * amount if not price_per_100g.empty else 0
        
        if menu_name not in ingredient_dict:
            ingredient_dict[menu_name] = []
        
        ingredient_dict[menu_name].append(Ingredient(name=ingredient_name, price=price, amount_g=amount))

    # menu_objects 생성
    menu_objects = {}
    for _, row in menu_nutri_df.iterrows():
        menu_name = row['Menu']
        nutrients = {
            'energy_kcal': row['energy_kcal'],
            'carbohydrate_g': row['carbohydrate_g'],
            'protein_g': row['protein_g'],
            'fat_g': row['fat_g'],
            'Ca_mg': row['Ca_mg']
        }
        ingredients = ingredient_dict.get(menu_name, [])
        category = menu_categories.get(menu_name, "Unknown")
        
        menu = Menu(name=menu_name, nutrients=nutrients, ingredients=ingredients, category=category)
        menu_objects[menu_name] = menu

    # Diet 생성 
    meals = []
    for _, row in diet_df.iterrows():
        meal_menus = row['Menus'].split(',')
        meal_menu_objects = [menu_objects[menu_name.strip()] for menu_name in meal_menus if menu_name.strip() in menu_objects]
        meals.append(Meal(meal_menu_objects, str(row['Day']), row['MealType']))

    return Diet(meals)

def load_all_menus(menu_db_path: str, ingre_db_path: str) -> List[Menu]:
    menu_ingre_df = pd.read_excel(menu_db_path, sheet_name='ingredient')
    menu_nutri_df = pd.read_excel(menu_db_path, sheet_name='nutrient')
    menu_cat_df = pd.read_excel(menu_db_path, sheet_name='category')
    ingre_price_df = pd.read_excel(ingre_db_path)

    # 카테고리 정보를 딕셔너리로 변환
    menu_categories = dict(zip(menu_cat_df['Menu'], menu_cat_df['Category']))

    # ingredient_dict 생성
    ingredient_dict = {}
    for _, row in menu_ingre_df.iterrows():
        menu_name = row['Menu']
        ingredient_name = row['Ingredient']
        amount = row['Amount_g']
        
        price_per_100g = ingre_price_df[ingre_price_df['Ingredient'] == ingredient_name]['Price']
        price = (price_per_100g.values[0] / 100) * amount if not price_per_100g.empty else 0
        
        if menu_name not in ingredient_dict:
            ingredient_dict[menu_name] = []
        
        ingredient_dict[menu_name].append(Ingredient(name=ingredient_name, price=price, amount_g=amount))

    # menu_objects 생성
    all_menus = []
    for _, row in menu_nutri_df.iterrows():
        menu_name = row['Menu']
        nutrients = {
            'energy_kcal': row['energy_kcal'],
            'carbohydrate_g': row['carbohydrate_g'],
            'protein_g': row['protein_g'],
            'fat_g': row['fat_g'],
            'Ca_mg': row['Ca_mg']
        }
        ingredients = ingredient_dict.get(menu_name, [])
        category = menu_categories.get(menu_name, "Unknown")
        
        menu = Menu(name=menu_name, nutrients=nutrients, ingredients=ingredients, category=category)
        all_menus.append(menu)

    return all_menus

def create_nutrient_constraints():
    min_values = {
        'energy_kcal': 1500,
        'carbohydrate_g': 230,
        'protein_g': 55,
        'fat_g': 50,
        'Ca_mg': 700
    }
    max_values = {
        'energy_kcal': 2200,
        'carbohydrate_g': 300,
        'protein_g': 100,
        'fat_g': 80,
        'Ca_mg': 2500
    }
    weights = {
        'energy_kcal': 1.0,
        'carbohydrate_g': 0.8,
        'protein_g': 1.2,
        'fat_g': 0.9,
        'Ca_mg': 0.7
    }
    
    return NutrientConstraints(min_values=min_values, max_values=max_values, weights=weights)

def load_sample_file(sample_path):
    sample = pd.read_excel(sample_path).iloc[4:22, 2:8]

    def extract_main_dish(dish):
        return dish.split('/')[0] if '/' in dish else dish

    result = []

    meal_types = ['Breakfast', 'Lunch', 'Dinner']
    for day in range(6):
        for idx, meal_type in enumerate(meal_types):
            start_idx = idx * 6
            menu_items = sample.iloc[start_idx:start_idx + 6, day].tolist()
            menu_items[0] = extract_main_dish(menu_items[0])
            menu_str = ', '.join(menu_items)
            result.append({
                'Day': day + 1,
                'MealType': meal_type,
                'Menus': menu_str
            })

    df_result = pd.DataFrame(result)

    book = load_workbook(sample_path)
    new_sheet_name = 'sample'
    sheet_number = 1
    while new_sheet_name in book.sheetnames:
        sheet_number += 1
        new_sheet_name = f'sample_{sheet_number}'

    with pd.ExcelWriter(sample_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        df_result.to_excel(writer, sheet_name=new_sheet_name, index=False)